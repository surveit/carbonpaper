"""profile_file — what a stored file holds, read before any schema has been declared for
it. The rows are real Q1 2026 Senate LDA filings, so what the profile says about types is
a claim about a file someone actually has rather than about one written to make a point."""
from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
import pytest

from app.services import workspace
from app.services.errors import FileNotStoredError
from app.services.uploads import save_upload
from app.tools import shared

# Four filings by one registrant, verbatim from the quarterly export. `income` is what
# makes the file interesting: the sheet stores it as characters ("40000.00"), and every
# value converts, so it is the case where the characters and the number both matter.
FILINGS = (
    b"client,registrant,income,issue_codes\n"
    b"O'GRADY PEYTON INTERNATIONAL (USA),MORRISON PUBLIC AFFAIRS GROUP,40000.00,Immigration\n"
    b"AMERICAN HOSPITAL ASSOCIATION,MORRISON PUBLIC AFFAIRS GROUP,20000.00,Immigration\n"
    b"AMERICAN HOSPITAL ASSOCIATION,MORRISON PUBLIC AFFAIRS GROUP,20000.00,Immigration\n"
    b"AMERICAN IMMIGRANT INVESTOR ALLIANCE,MORRISON PUBLIC AFFAIRS GROUP,,Immigration\n"
)


@pytest.fixture
def project(tmp_path, monkeypatch):
    (tmp_path / "demo").mkdir(parents=True)
    workspace.set_projects_dir(tmp_path)
    monkeypatch.setenv("CARBON_PAPER_FILES_ROOT", str(tmp_path / "files"))
    return tmp_path / "demo"


def store(name: str = "lda_q1.csv", body: bytes = FILINGS) -> str:
    return save_upload(name, io.BytesIO(body), "demo").sha256


def column_of(profile, name: str):
    return next(column for column in profile.columns if column.column == name)


def test_it_profiles_every_column_without_being_told_the_columns(project):
    profile = shared.profile_file("demo", store())
    # The whole reason the tool exists: the caller does not know the header row yet.
    assert [column.column for column in profile.columns] == [
        "client", "registrant", "income", "issue_codes"]
    assert profile.row_count == 4



def test_a_repeated_value_carries_its_count(project):
    client = column_of(shared.profile_file("demo", store()), "client")
    assert client.values[0].value == "AMERICAN HOSPITAL ASSOCIATION"
    assert client.values[0].count == 2
    assert client.distinct_count == 3
    assert client.truncated is False


def test_a_blank_cell_counts_as_null_not_as_a_value(project):
    income = column_of(shared.profile_file("demo", store()), "income")
    assert income.null_count == 1
    assert [value.value for value in income.values] == ["20000.00", "40000.00"]


def test_money_keeps_its_characters_and_still_reports_a_range(project):
    """Both halves of the schema decision arrive together, off one read."""
    income = column_of(shared.profile_file("demo", store()), "income")
    assert [value.value for value in income.values] == ["20000.00", "40000.00"]
    assert income.value_range is not None
    assert (income.value_range.min, income.value_range.max) == (20000.0, 40000.0)


def test_a_zero_padded_id_survives_the_read(project):
    """A plain read makes "002" the integer 2; a schema declared off that loses the padding."""
    padded = b"filing_id,client\n002,ACME\n010,BETA\n"
    filing_id = column_of(shared.profile_file("demo", store("padded.csv", padded)), "filing_id")
    assert [value.value for value in filing_id.values] == ["002", "010"]
    # Read plainly, this same column is 2 and 10 — which is what the range reports, so
    # the profile shows the loss rather than hiding it behind either answer alone.
    assert pd.read_csv(io.BytesIO(padded))["filing_id"].tolist() == [2, 10]
    assert filing_id.value_range is not None
    assert (filing_id.value_range.min, filing_id.value_range.max) == (2.0, 10.0)


def test_a_column_that_is_not_all_numbers_reports_no_range(project):
    issue = column_of(shared.profile_file("demo", store()), "issue_codes")
    assert issue.value_range is None


def test_naming_a_column_profiles_only_that_one(project):
    profile = shared.profile_file("demo", store(), ["registrant"])
    assert [column.column for column in profile.columns] == ["registrant"]


def test_a_column_the_file_does_not_hold_is_loud_and_names_what_it_does(project):
    with pytest.raises(ValueError) as caught:
        shared.profile_file("demo", store(), ["Registrant Name"])
    assert "client" in str(caught.value) and "issue_codes" in str(caught.value)


def test_max_values_cuts_the_list_and_says_so(project):
    client = column_of(shared.profile_file("demo", store(), None, 1), "client")
    assert len(client.values) == 1
    # The true count survives the cut, so nobody declares a 1-value enum off a prefix.
    assert client.distinct_count == 3
    assert client.truncated is True


def test_max_values_below_one_is_refused(project):
    with pytest.raises(ValueError, match="max_values"):
        shared.profile_file("demo", store(), None, 0)


def test_a_file_the_project_does_not_hold_is_refused(project):
    store()
    with pytest.raises(FileNotStoredError, match="has no file"):
        shared.profile_file("demo", "0" * 64)


def test_a_file_in_no_project_is_not_readable_until_one_takes_it(project, tmp_path):
    """Scoping is by (sha256, project): the bytes exist, but no project holds them yet."""
    loose = save_upload("dropped.csv", io.BytesIO(FILINGS), None).sha256
    with pytest.raises(FileNotStoredError, match="has no file"):
        shared.profile_file("demo", loose)
    shared.move_file_to_project("demo", loose)
    assert shared.profile_file("demo", loose).row_count == 4


def test_a_record_whose_bytes_are_gone_says_so_rather_than_failing_on_the_read(project):
    from app.services.uploads import files_root
    sha = store()
    (files_root() / sha / "lda_q1.csv").unlink()
    with pytest.raises(FileNotStoredError, match="not on disk"):
        shared.profile_file("demo", sha)


def test_a_missing_project_is_refused_before_the_file_is_looked_up(project):
    with pytest.raises(ValueError, match="no project"):
        shared.profile_file("nope", store())


# ── Workbooks: many sheets, and a table that does not start at A1 ────────────

def _workbook(tmp_path):
    """Sheet 2 carries a title row and an indented table — the shape survey_workbook is for."""
    import openpyxl
    wb = openpyxl.Workbook()
    plain = wb.active
    plain.title = "Filings"
    for row in [["client", "income"], ["COMCAST CORPORATION", "40000.00"]]:
        plain.append(row)
    offset = wb.create_sheet("Q1 Summary")
    offset.append(["LOBBYING DISCLOSURE — Q1 2026"])
    offset.append([])
    offset.append([None, "registrant", "filings"])
    offset.append([None, "CORNERSTONE GOVERNMENT AFFAIRS", "392"])
    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


def store_workbook(tmp_path) -> str:
    with _workbook(tmp_path).open("rb") as handle:
        return save_upload("book.xlsx", handle, "demo").sha256


def test_the_survey_names_every_sheet(project, tmp_path):
    sheets = shared.survey_workbook("demo", store_workbook(tmp_path))
    assert [sheet.name for sheet in sheets] == ["Filings", "Q1 Summary"]


def test_the_survey_shows_where_the_table_really_starts(project, tmp_path):
    """The cells are the answer: the caller reads header_row and first_column off them."""
    offset = shared.survey_workbook("demo", store_workbook(tmp_path))[1]
    assert offset.cells[0][0] == "LOBBYING DISCLOSURE — Q1 2026"
    assert offset.cells[1][0] is None
    # The indices ARE the two arguments: row 2, column 1.
    assert offset.cells[2][:3] == [None, "registrant", "filings"]
    assert offset.first_row == 0


def test_profiling_that_sheet_with_the_defaults_gets_junk(project, tmp_path):
    """Why the survey exists: the title row becomes the header and the table is lost."""
    sha = store_workbook(tmp_path)
    columns = shared.profile_file("demo", sha, sheet_name="Q1 Summary").columns
    assert [c.column for c in columns] == [
        "LOBBYING DISCLOSURE — Q1 2026", "Unnamed: 1", "Unnamed: 2"]


def test_profiling_that_sheet_with_the_surveyed_offsets_gets_the_table(project, tmp_path):
    sha = store_workbook(tmp_path)
    profile = shared.profile_file(
        "demo", sha, sheet_name="Q1 Summary", header_row=2, first_column=1)
    assert [c.column for c in profile.columns] == ["registrant", "filings"]
    assert profile.row_count == 1
    assert profile.columns[0].values[0].value == "CORNERSTONE GOVERNMENT AFFAIRS"


def test_a_named_sheet_is_read_rather_than_the_first(project, tmp_path):
    profile = shared.profile_file("demo", store_workbook(tmp_path), sheet_name="Filings")
    assert [c.column for c in profile.columns] == ["client", "income"]


def test_surveying_a_csv_is_refused_and_says_what_to_call_instead(project):
    with pytest.raises(ValueError, match="no.*sheets"):
        shared.survey_workbook("demo", store())


def _long_preamble_workbook(tmp_path) -> Path:
    """Seven rows of letterhead before the header — longer than one window."""
    import openpyxl
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.title = "Extract"
    for line in ["Senate Office of Public Records", "", "Extract generated 2026-04-20",
                 "Coverage: 2026-01-01 to 2026-03-31", "Contact: records@senate.gov",
                 "", "NOTES"]:
        sheet.append([line])
    sheet.append(["registrant", "filings"])
    sheet.append(["CORNERSTONE GOVERNMENT AFFAIRS", "392"])
    path = tmp_path / "long.xlsx"
    wb.save(path)
    return path


def store_long(tmp_path) -> str:
    with _long_preamble_workbook(tmp_path).open("rb") as handle:
        return save_upload("long.xlsx", handle, "demo").sha256


def test_a_preamble_longer_than_the_window_hides_the_header(project, tmp_path):
    """The window is a window: it does not hunt for the table, so this shows only prose."""
    sheet = shared.survey_workbook("demo", store_long(tmp_path))[0]
    assert [row[0] for row in sheet.cells] == [
        "Senate Office of Public Records", None, "Extract generated 2026-04-20",
        "Coverage: 2026-01-01 to 2026-03-31", "Contact: records@senate.gov"]


def test_surveying_again_from_further_down_finds_it(project, tmp_path):
    sheet = shared.survey_workbook("demo", store_long(tmp_path), 5)[0]
    assert sheet.first_row == 5
    # cells[2] is sheet row 5 + 2 = 7, which is the header_row profile_file then takes.
    assert sheet.cells[2][:2] == ["registrant", "filings"]
    profile = shared.profile_file(
        "demo", store_long(tmp_path), sheet_name="Extract", header_row=7)
    assert [c.column for c in profile.columns] == ["registrant", "filings"]
    assert profile.row_count == 1


def test_a_negative_from_row_is_refused(project, tmp_path):
    with pytest.raises(ValueError, match="from_row"):
        shared.survey_workbook("demo", store_long(tmp_path), -1)
