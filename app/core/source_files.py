"""Reading a SOURCE file someone else wrote: the format vocabulary, what extension
names each one, and the one read that turns a path into a frame. Held below both the
runtime and the services so the two cannot drift on what a format means."""
from __future__ import annotations

import json
from collections.abc import Hashable, Mapping
from enum import Enum
from pathlib import Path
from typing import Any, NamedTuple

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from app.core.frames import (
    STR_TYPE_NAME,
    read_frame_file,
    read_source_csv,
    read_source_excel,
    read_source_json_lines,
)


class FileFormat(str, Enum):
    csv = "csv"
    tsv = "tsv"
    parquet = "parquet"
    json = "json"
    geojson = "geojson"
    xlsx = "xlsx"


# The extension each format is designated by on disk. `.jsonl` names the same
# reader as `.json` (both are read line-delimited).
_FORMAT_BY_SUFFIX: dict[str, FileFormat] = {
    ".csv": FileFormat.csv,
    ".tsv": FileFormat.tsv,
    ".parquet": FileFormat.parquet,
    ".json": FileFormat.json,
    ".jsonl": FileFormat.json,
    ".geojson": FileFormat.geojson,
    ".xlsx": FileFormat.xlsx,
}

# What pandas takes as `dtype=`: a pin per named column, one type for every column, or
# `False` (json only) for no inference at all. Which of the three a caller wants depends
# on whether it has a declared schema, so this layer takes the answer rather than one.
SourceDtype = Mapping[Hashable, Any] | type | bool | None


# app.core cannot import app.models' Column.type, so callers pass plain strings.
_TEXT_ON_DISK_TYPES = frozenset({STR_TYPE_NAME, "json", "date", "datetime"})


def text_on_disk_columns(column_types: Mapping[str, str], fmt: FileFormat) -> list[str]:
    """Columns to pin to `str` so a type pandas infers (a zero-padded id) survives."""
    if fmt in (FileFormat.csv, FileFormat.tsv, FileFormat.xlsx):
        return [name for name, type_ in column_types.items()
                if type_ in _TEXT_ON_DISK_TYPES or type_.startswith("list[")]
    if fmt == FileFormat.json:
        return [name for name, type_ in column_types.items() if type_ == STR_TYPE_NAME]
    return []


def find_file_format(path: str) -> FileFormat | None:
    """None means no reader here holds this extension — a png, a pdf, a zip."""
    return _FORMAT_BY_SUFFIX.get(Path(path).suffix.lower())


def resolve_file_format(path: str) -> FileFormat:
    suffix = Path(path).suffix.lower()
    fmt = _FORMAT_BY_SUFFIX.get(suffix)
    if fmt is None:
        raise ValueError(
            f"cannot tell what format {path!r} holds: extension "
            f"{suffix or '(none)'} is not one of {sorted(_FORMAT_BY_SUFFIX)}"
        )
    return fmt


def read_source_file(
    path: Path,
    fmt: FileFormat,
    *,
    dtype: SourceDtype = None,
    sheet_name: str | int = 0,
    header_row: int = 0,
    first_column: int = 0,
    source_row_column: str | None = None,
) -> pd.DataFrame:
    """`dtype` reaches only the formats pandas infers; parquet and geojson carry types."""
    if fmt == FileFormat.csv:
        return read_source_csv(path, dtype=dtype)
    if fmt == FileFormat.tsv:
        return read_source_csv(path, dtype=dtype, delimiter="\t")
    if fmt == FileFormat.parquet:
        return read_frame_file(path)
    if fmt == FileFormat.json:
        return read_source_json_lines(path, dtype=dtype)
    if fmt == FileFormat.geojson:
        return read_source_geojson(path)
    if fmt == FileFormat.xlsx:
        return _read_xlsx(path, dtype=dtype, sheet_name=sheet_name, header_row=header_row,
                          first_column=first_column, source_row_column=source_row_column)
    raise ValueError(f"Unsupported file format: {fmt}")


class SheetSurvey(NamedTuple):
    name: str
    row_count: int
    column_count: int
    # Where `cells` starts, so a caller reading an index off it lands on the right row
    # after paging past a preamble: sheet row = first_row + index into `cells`.
    first_row: int
    # A fixed window of the sheet as it sits, no header chosen and nothing skipped. The
    # VALUES are what separate a title from a header from a first data row; the indices
    # into them are the header_row/first_column a read then takes.
    cells: list[list[str | None]]


def survey_xlsx_sheets(
    path: Path, *, from_row: int = 0, rows: int = 5, columns: int = 8,
) -> list[SheetSurvey]:
    """Reads no data: openpyxl's read-only mode streams the window and the dimensions."""
    if from_row < 0:
        raise ValueError(f"from_row must be at least 0, got {from_row}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [_survey_one_sheet(workbook[name], name, from_row, rows, columns)
                for name in workbook.sheetnames]
    finally:
        workbook.close()


def _survey_one_sheet(
    sheet: Worksheet, name: str, from_row: int, rows: int, columns: int,
) -> SheetSurvey:
    window = [
        [None if cell.value is None else str(cell.value) for cell in row]
        # openpyxl rows are 1-based; from_row is the 0-based index header_row also uses.
        for row in sheet.iter_rows(
            min_row=from_row + 1, max_row=from_row + rows, max_col=columns)
    ]
    return SheetSurvey(
        name=name,
        # openpyxl reports the sheet's declared extent, which counts a trailing styled
        # but empty row, so this is an upper bound rather than the row count a read gives.
        row_count=sheet.max_row or 0,
        column_count=sheet.max_column or 0,
        first_row=from_row,
        cells=window,
    )


def read_source_geojson(path: Path) -> pd.DataFrame:
    """One row per feature, its properties flat; a Point also gets `lon`/`lat`."""
    geo = json.loads(path.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for feat in geo.get("features", []):
        props = dict(feat.get("properties") or {})
        geom = feat.get("geometry") or {}
        if geom.get("type") == "Point":
            coords = geom.get("coordinates") or [None, None]
            props.setdefault("lon", coords[0])
            props.setdefault("lat", coords[1])
        rows.append(props)
    return pd.DataFrame(rows)


def _read_xlsx(
    path: Path, *, dtype: SourceDtype, sheet_name: str | int, header_row: int,
    first_column: int, source_row_column: str | None,
) -> pd.DataFrame:
    # header_row/first_column are 0-based indices into the sheet as Excel shows it.
    frame = read_source_excel(
        path, sheet_name=sheet_name, header_row=header_row, dtype=dtype
    )
    if first_column:
        _refuse_first_column_out_of_range(first_column, frame, path, sheet_name)
        frame = frame.iloc[:, first_column:].copy()
    if source_row_column:
        _add_source_row_column(frame, source_row_column, header_row)
    return frame


def _add_source_row_column(frame: pd.DataFrame, column: str, header_row: int) -> None:
    # Sheet rows are 1-based and the data starts one row after the header, hence + 2.
    if column in frame.columns:
        raise ValueError(f"source_row_column '{column}' collides with an existing column")
    frame[column] = frame.index + header_row + 2


def _refuse_first_column_out_of_range(
    first_column: int, frame: pd.DataFrame, path: Path, sheet: Any
) -> None:
    if first_column < 0 or first_column >= len(frame.columns):
        raise ValueError(
            f"first_column={first_column} is out of range for {path.name} "
            f"sheet {sheet!r}, which has {len(frame.columns)} columns"
        )
